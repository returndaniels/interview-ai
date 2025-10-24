from openpyxl import load_workbook
from fastapi import UploadFile
from ..utils import sanitize_sql_name, get_sql_type
from ..database import get_db_connection, get_db_cursor, close_db_connection

import pandas as pd
import io

def get_tables():
    connection = get_db_connection()
    cursor = connection.cursor()
    
    cursor.execute("SHOW TABLES")
    all_tables = [table[0] for table in cursor.fetchall()]
    
    # Filtra apenas tabelas com prefixo datasheet_
    datasheet_tables = [t for t in all_tables if t.startswith('datasheet_')]
    
    cursor.close()
    connection.close()
    return datasheet_tables

def read_excel_in_chunks(file, batch_size=1000):
    """Lê um arquivo Excel em chunks de DataFrames."""
    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows = []
    headers = [cell.value for cell in next(ws.rows)]  # primeira linha
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # já lemos o cabeçalho
        rows.append(row)
        if len(rows) == batch_size:
            yield pd.DataFrame(rows, columns=headers)
            rows = []
    if rows:
        yield pd.DataFrame(rows, columns=headers)


def create_table_from_dataframe(cursor, table_name, df):
    """ Cria uma tabela no banco de dados com base no DataFrame fornecido."""
    columns = df.columns
    types = df.dtypes
    sanitized_columns = [sanitize_sql_name(col) for col in columns]
    col_defs = ", ".join([f"`{col}` {get_sql_type(dtype)}" for col, dtype in zip(sanitized_columns, types)])
    create_table_query = f"CREATE TABLE IF NOT EXISTS `datasheet_{table_name}` ({col_defs});"
    cursor.execute(create_table_query)


async def import_excel_to_database(upload_file: UploadFile, table_name: str):
    """
    Importa um arquivo Excel (UploadFile do FastAPI) para o banco de dados MariaDB.
    
    Args:
        upload_file: Objeto UploadFile do FastAPI contendo o arquivo Excel
        table_name: Nome da tabela a ser criada
    
    Returns:
        dict: Informações sobre a importação
    """
    connection = None
    cursor = None
    
    try:
        # Lê o conteúdo do arquivo em memória
        contents = await upload_file.read()
        file_like = io.BytesIO(contents)
        
        # Conecta ao banco
        connection = get_db_connection()
        cursor = get_db_cursor(connection)
        
        total_rows = 0
        first_chunk = True
        
        # Processa o arquivo em chunks
        for chunk_df in read_excel_in_chunks(file_like):
            # Cria a tabela apenas no primeiro chunk
            if first_chunk:
                create_table_from_dataframe(cursor, table_name, chunk_df)
                connection.commit()
                first_chunk = False
            
            # Insere os dados
            sanitized_columns = [sanitize_sql_name(col) for col in chunk_df.columns]
            placeholders = ", ".join(["%s"] * len(sanitized_columns))
            columns_str = ", ".join([f"`{col}`" for col in sanitized_columns])
            prefixed_table_name = table_name if table_name.startswith('datasheet_') else f"datasheet_{table_name}"
            
            insert_query = f"INSERT INTO `{prefixed_table_name}` ({columns_str}) VALUES ({placeholders})"
            
            # Converte DataFrame para lista de tuplas
            data = [tuple(row) for row in chunk_df.values]
            cursor.executemany(insert_query, data)
            connection.commit()
            
            total_rows += len(chunk_df)

        _dict = {
            "success": True,
            "table_name": f"{prefixed_table_name}",
            "rows_imported": total_rows,
            "filename": upload_file.filename
        }
        message = f"Importação concluída: {total_rows} linhas inseridas na tabela '{prefixed_table_name}'"
        
        return _dict, message
        
    except Exception as e:
        if connection:
            connection.rollback()
        raise Exception(f"Erro ao importar Excel para banco: {str(e)}")
        
    finally:
        close_db_connection(connection, cursor)


def get_table_data_paginated(
    table_name: str,
    page: int = 1,
    page_size: int = 50,
    search: str = None,
    sort_by: str = None,
    sort_order: str = "asc"
):
    """
    Retorna os dados de uma tabela com paginação e filtros
    
    Args:
        table_name: Nome da tabela (deve começar com datasheet_)
        page: Número da página (inicia em 1)
        page_size: Quantidade de registros por página (máximo 100)
        search: Termo de busca (busca em todas as colunas de texto)
        sort_by: Nome da coluna para ordenação
        sort_order: Ordem de classificação (asc ou desc)
    
    Returns:
        Tupla (success: bool, result: dict)
        
        Em caso de sucesso:
        {
            "table_name": "datasheet_vendas",
            "data": [...],
            "pagination": {
                "page": 1,
                "page_size": 50,
                "total_records": 1000,
                "total_pages": 20
            },
            "columns": ["id", "nome", "valor"]
        }
        
        Em caso de erro:
        {
            "error": "mensagem de erro",
            "error_type": "validation|not_found|database"
        }
    """
    
    # Validação: tabela deve começar com datasheet_
    if not table_name.startswith("datasheet_"):
        return False, {
            "error": "Acesso negado. Apenas tabelas com prefixo 'datasheet_' são permitidas.",
            "error_type": "validation"
        }
    
    # Sanitiza o nome da tabela
    safe_table_name = sanitize_sql_name(table_name)
    
    # Validação de paginação
    if page < 1:
        page = 1
    if page_size < 1:
        page_size = 50
    if page_size > 100:
        page_size = 100
    
    # Validação de sort_order
    if sort_order.lower() not in ["asc", "desc"]:
        sort_order = "asc"
    
    try:
        cursor = get_db_cursor()
        
        # Verifica se a tabela existe
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() 
            AND table_name = %s
        """, (table_name,))
        
        if cursor.fetchone()["count"] == 0:
            cursor.close()
            return False, {
                "error": f"Tabela '{table_name}' não encontrada",
                "error_type": "not_found"
            }
        
        # Obtém as colunas da tabela
        cursor.execute(f"DESCRIBE `{safe_table_name}`")
        columns = [col["Field"] for col in cursor.fetchall()]
        
        # Monta a query base
        base_query = f"FROM `{safe_table_name}`"
        where_clause = ""
        params = []
        
        # Adiciona filtro de busca se fornecido
        if search:
            # Busca em todas as colunas de texto
            search_conditions = []
            for col in columns:
                search_conditions.append(f"`{sanitize_sql_name(col)}` LIKE %s")
            
            where_clause = f" WHERE ({' OR '.join(search_conditions)})"
            params = [f"%{search}%"] * len(columns)
        
        # Conta total de registros
        count_query = f"SELECT COUNT(*) as total {base_query} {where_clause}"
        cursor.execute(count_query, params)
        total_records = cursor.fetchone()["total"]
        
        # Calcula total de páginas
        total_pages = (total_records + page_size - 1) // page_size
        
        # Monta query de dados com paginação
        offset = (page - 1) * page_size
        
        order_clause = ""
        if sort_by and sort_by in columns:
            safe_sort_by = sanitize_sql_name(sort_by)
            order_clause = f" ORDER BY `{safe_sort_by}` {sort_order.upper()}"
        
        data_query = f"SELECT * {base_query} {where_clause} {order_clause} LIMIT %s OFFSET %s"
        cursor.execute(data_query, params + [page_size, offset])
        
        data = cursor.fetchall()
        cursor.close()
        
        return True, {
            "table_name": table_name,
            "data": data,
            "pagination": {
                "page": page,
                "page_size": page_size,
                "total_records": total_records,
                "total_pages": total_pages
            },
            "columns": columns
        }
        
    except Exception as e:
        return False, {
            "error": f"Erro ao buscar dados da tabela: {str(e)}",
            "error_type": "database"
        }
